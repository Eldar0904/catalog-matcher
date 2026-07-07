"""
Export final (possibly user-overridden) matches to .xlsx file(s) via
openpyxl (through pandas for simplicity).

Modes:
  - full export: every internal item, matched or not (one file)
  - best-matches export: only items whose selected match's confidence
    score is at or above a threshold (e.g. 0.8 for >=80%), one file
  - batched best-matches export: same filter as above, but split into
    multiple .xlsx files of up to `batch_size` rows each, bundled into
    a single .zip for download (useful when the buyer/portal only
    accepts uploads of a limited row count)
"""
import os
import zipfile
from datetime import datetime
from typing import Optional, List

import pandas as pd
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill,
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import settings
from app.models.db_models import InternalItem, MatchResult, CatalogProduct

# ── Palette (matches main dashboard) ──────────────────────────────────────────
_NAVY        = "1C3557"
_NAVY_LIGHT  = "254A72"
_WHITE       = "FFFFFF"
_ROW_ALT     = "F1F5F9"   # slate-100
_ROW_BASE    = "FFFFFF"
_GREEN_BG    = "DCFCE7"
_GREEN_FG    = "15803D"
_AMBER_BG    = "FEF3C7"
_AMBER_FG    = "B45309"
_RED_BG      = "FEE2E2"
_RED_FG      = "DC2626"
_BORDER_CLR  = "E2E8F0"
_NO_MATCH_BG = "F8FAFC"
_NO_MATCH_FG = "94A3B8"

def _thin_border(color=_BORDER_CLR):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _conf_colors(score):
    """Return (bg_hex, fg_hex) for a confidence 0-1 score."""
    if score is None:
        return _NO_MATCH_BG, _NO_MATCH_FG
    if score >= 0.6:
        return _GREEN_BG, _GREEN_FG
    if score >= 0.3:
        return _AMBER_BG, _AMBER_FG
    return _RED_BG, _RED_FG


def _build_rows(db: Session, min_confidence: Optional[float] = None):
    items = db.query(InternalItem).all()

    rows = []
    for item in items:
        selected = (
            db.query(MatchResult)
            .filter(MatchResult.item_id == item.id, MatchResult.is_selected == 1)
            .first()
        )
        product = selected.catalog_product if selected else None

        if min_confidence is not None:
            # Best-matches export: skip items with no match or below threshold
            if not selected or selected.confidence_score < min_confidence:
                continue

        rows.append({
            "Item Code": item.item_code,
            "Item Name": item.item_name,
            "Item Description": item.description,
            "Quantity": item.quantity,
            "Matched Government Code": product.code if product else None,
            "Matched Product Name": product.name if product else None,
            "Matched Brand": product.brand if product else None,
            "Matched Model": product.model if product else None,
            "Matched Price": product.price if product else None,
            "Confidence Score": round(selected.confidence_score, 3) if selected else None,
            "Explanation": selected.explanation if selected else None,
            "Manual Override": "Yes" if (selected and selected.is_manual_override) else "No",
            "Match Status": "Matched" if product else "No Match",
        })

    return rows


def _write_xlsx(df: pd.DataFrame, filepath: str, sheet_name: str):
    # Column widths (characters) — override per column name, fallback = auto
    COL_WIDTHS = {
        "Item Code":               14,
        "Item Name":               38,
        "Item Description":        34,
        "Quantity":                10,
        "Matched Government Code": 22,
        "Matched Product Name":    40,
        "Matched Brand":           18,
        "Matched Model":           18,
        "Matched Price":           14,
        "Confidence Score":        16,
        "Explanation":             36,
        "Manual Override":         14,
        "Match Status":            14,
    }

    # Which column index (1-based) holds the confidence score
    conf_col_idx = None
    if "Confidence Score" in df.columns:
        conf_col_idx = list(df.columns).index("Confidence Score") + 1

    status_col_idx = None
    if "Match Status" in df.columns:
        status_col_idx = list(df.columns).index("Match Status") + 1

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        n_cols = len(df.columns)

        # ── Header row (row 1) ─────────────────────────────────────────────
        hdr_fill   = PatternFill("solid", fgColor=_NAVY)
        hdr_font   = Font(name="Calibri", bold=True, color=_WHITE, size=10)
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
        hdr_border = _thin_border(_NAVY_LIGHT)

        ws.row_dimensions[1].height = 26

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.alignment = hdr_align
            cell.border = hdr_border

        # ── Data rows ──────────────────────────────────────────────────────
        base_font   = Font(name="Calibri", size=9.5)
        base_align  = Alignment(vertical="center", wrap_text=True)
        border      = _thin_border()

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            is_alt = (row_idx % 2 == 0)
            row_bg = _ROW_ALT if is_alt else _ROW_BASE
            ws.row_dimensions[row_idx].height = 36

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font      = base_font
                cell.alignment = base_align
                cell.border    = border

                # Confidence score cell — color-code by tier
                if col_idx == conf_col_idx:
                    score = cell.value
                    num_score = score if isinstance(score, (int, float)) else None
                    bg, fg = _conf_colors(num_score)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(name="Calibri", size=9.5, bold=True, color=fg)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Display as percentage string
                    if num_score is not None:
                        cell.value = f"{round(num_score * 100)}%"
                    continue

                # Match status cell
                if col_idx == status_col_idx:
                    val = cell.value
                    if val == "Matched":
                        cell.fill = PatternFill("solid", fgColor=_GREEN_BG)
                        cell.font = Font(name="Calibri", size=9.5, bold=True, color=_GREEN_FG)
                    elif val == "No Match":
                        cell.fill = PatternFill("solid", fgColor=_NO_MATCH_BG)
                        cell.font = Font(name="Calibri", size=9.5, color=_NO_MATCH_FG)
                    else:
                        cell.fill = PatternFill("solid", fgColor=row_bg)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    continue

                cell.fill = PatternFill("solid", fgColor=row_bg)

        # ── Column widths ──────────────────────────────────────────────────
        for col_idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(col_idx)
            if col_name in COL_WIDTHS:
                ws.column_dimensions[letter].width = COL_WIDTHS[col_name]
            else:
                # auto
                max_len = max(
                    [len(str(col_name))] +
                    [len(str(v)) for v in df[col_name].astype(str)]
                ) if len(df) else len(str(col_name))
                ws.column_dimensions[letter].width = min(max_len + 2, 50)

        # ── Freeze header + first column ───────────────────────────────────
        ws.freeze_panes = "B2"

        # ── Summary row at bottom ──────────────────────────────────────────
        if len(df) > 0 and status_col_idx:
            summary_row = len(df) + 2
            ws.row_dimensions[summary_row].height = 20
            total_label_cell = ws.cell(row=summary_row, column=1)
            total_label_cell.value     = f"Итого: {len(df)} позиций"
            total_label_cell.font      = Font(name="Calibri", size=9.5, bold=True, color=_NAVY)
            total_label_cell.alignment = Alignment(vertical="center")

            # Count by status
            if "Match Status" in df.columns:
                matched_count  = (df["Match Status"] == "Matched").sum()
                no_match_count = (df["Match Status"] == "No Match").sum()
                summary_cell = ws.cell(row=summary_row, column=status_col_idx)
                summary_cell.value     = f"✓ {matched_count} подобрано  ✗ {no_match_count} без совпадений"
                summary_cell.font      = Font(name="Calibri", size=9.5, bold=True, color=_NAVY)
                summary_cell.alignment = Alignment(horizontal="center", vertical="center")


def export_results(db: Session, min_confidence: Optional[float] = None) -> str:
    """
    Export matching results to a single .xlsx file.

    min_confidence=None -> full export (all items, matched or not)
    min_confidence=0.8   -> only items with a selected match >= 80% confidence
    """
    rows = _build_rows(db, min_confidence)
    df = pd.DataFrame(rows)

    os.makedirs(settings.exports_dir, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if min_confidence is not None:
        pct = int(round(min_confidence * 100))
        filename = f"Best_Matches_{pct}pct_{timestamp}.xlsx"
        sheet_name = "Best Matches"
    else:
        filename = f"Export_{timestamp}.xlsx"
        sheet_name = "Export"

    filepath = os.path.join(settings.exports_dir, filename)
    _write_xlsx(df, filepath, sheet_name)
    return filepath


def export_results_batched(
    db: Session,
    min_confidence: Optional[float] = None,
    batch_size: int = 100,
) -> str:
    """
    Same filtering as export_results, but splits the rows into multiple
    .xlsx files of up to `batch_size` rows each and bundles them into one
    .zip for download. Returns the path to the .zip.
    """
    if batch_size < 1:
        raise ValueError("batch_size must be >= 1")

    rows = _build_rows(db, min_confidence)

    os.makedirs(settings.exports_dir, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    pct_label = f"{int(round(min_confidence * 100))}pct_" if min_confidence is not None else ""
    zip_name = f"Best_Matches_{pct_label}batches_of_{batch_size}_{timestamp}.zip"
    zip_path = os.path.join(settings.exports_dir, zip_name)

    batch_paths: List[str] = []

    if not rows:
        # Still produce a zip with a single empty batch so the response is
        # never confusingly missing — makes "no matches found" explicit.
        df = pd.DataFrame(rows)
        batch_path = os.path.join(settings.exports_dir, f"_tmp_batch_1_{timestamp}.xlsx")
        _write_xlsx(df, batch_path, "Best Matches")
        batch_paths.append(batch_path)
    else:
        total_batches = (len(rows) + batch_size - 1) // batch_size
        for i in range(total_batches):
            batch_rows = rows[i * batch_size:(i + 1) * batch_size]
            df = pd.DataFrame(batch_rows)
            batch_path = os.path.join(settings.exports_dir, f"_tmp_batch_{i + 1}_{timestamp}.xlsx")
            _write_xlsx(df, batch_path, "Best Matches")
            batch_paths.append(batch_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, batch_path in enumerate(batch_paths, start=1):
            arcname = f"Best_Matches_batch_{idx}.xlsx"
            zf.write(batch_path, arcname=arcname)

    # Clean up the individual batch files now that they're zipped
    for batch_path in batch_paths:
        try:
            os.remove(batch_path)
        except OSError:
            pass

    return zip_path
